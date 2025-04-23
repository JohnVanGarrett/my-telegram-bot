import os
import time
import datetime
import logging
from openpyxl import Workbook, load_workbook
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ChatMemberHandler,
    ContextTypes,
)

# ====== 1. УСТАНАВЛИВАЕМ ЧАСОВОЙ ПОЯС МСК ======
os.environ['TZ'] = 'Europe/Moscow'
time.tzset()

# ====== 2. ЛОГИРОВАНИЕ ======
logging.basicConfig(
    filename='bot.log',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s'
)

# ====== 3. ПУТЬ К EXCEL‑ФАЙЛУ ======
XLSX_FILE = 'visitors.xlsx'

def init_excel():
    if not os.path.exists(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            'timestamp',
            'user_id',
            'username',
            'action',
            'clicked_subscribed',
            'blocked_bot',
            'duration_seconds'
        ])
        wb.save(XLSX_FILE)

def log_to_excel(timestamp, user_id, username, action, clicked_sub, blocked_bot, duration):
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    ws.append([timestamp, user_id, username, action, clicked_sub, blocked_bot, duration])
    wb.save(XLSX_FILE)

# Инициализируем Excel‑файл
init_excel()

# ====== КОНФИГ БОТА ======
TOKEN       = os.environ['TOKEN']
CHANNEL     = "@photo_therapy_art"
ARTICLE_URL = "https://taplink.cc/yapropala"

WELCOME_TEXT = (
    "Привет 😍\n"
    "Рад что ты здесь!\n\n"
    "Всё, что сейчас надо сделать – подписаться на мой канал:\n"
    "https://t.me/photo_therapy_art\n"
    "https://t.me/photo_therapy_art\n"
    "https://t.me/photo_therapy_art\n\n"
    "А после подписки жми «Я УЖЕ ПОДПИСАН(А)»,\n"
    "и я выдам тебе статью, в которой раскрою главный секрет —\n"
    "почему ты теряешь контакт с собой."
)

MSG_OK = (
    "Супер! Вижу твою подписку 😃\n\n"
    "Поздравляю! Вот твоя статья,переходи по ссылке\n\n"
    f"{ARTICLE_URL}\n"
    f"{ARTICLE_URL}\n"
    f"{ARTICLE_URL}\n\n"
    "\nОна ответит тебе на самый главный вопрос - как вернуть контакт с собой. ПЕРЕХОДИ И ЧИТАЙ!"
)

MSG_FAIL = (
    "Хмм…\n"
    "Не вижу тебя среди подписчиков.\n"
    "Подпишись и жми «Я УЖЕ ПОДПИСАН(А)» ещё раз:\n"
    "https://t.me/photo_therapy_art\n"
    "https://t.me/photo_therapy_art\n"
    "https://t.me/photo_therapy_art"
)

def now_moscow():
    return (datetime.datetime.utcnow() + datetime.timedelta(hours=3))\
           .isoformat(sep=' ')

# Обработчик команды /start
async def start(update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    uid  = user.id
    uname = user.username or str(uid)
    ts = now_moscow()

    logging.info(f"/start by user_id={uid} username={uname}")
    log_to_excel(ts, uid, uname, 'start', False, False, 0)

    kb = [[InlineKeyboardButton("Я УЖЕ ПОДПИСАН(А)", callback_data="check_sub")]]
    await update.message.reply_text(WELCOME_TEXT, reply_markup=InlineKeyboardMarkup(kb))

# Обработчик нажатия «Я УЖЕ ПОДПИСАН(А)»
async def check_subscribed(update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    user = update.callback_query.from_user
    uid  = user.id
    uname = user.username or str(uid)
    ts = now_moscow()

    logging.info(f"check_sub by user_id={uid} username={uname}")
    log_to_excel(ts, uid, uname, 'check_sub', True, False, 0)

    try:
        member = await context.bot.get_chat_member(CHANNEL, uid)
        status = member.status
    except Exception as e:
        logging.error(f"Subscription check error for {uid}: {e}")
        await update.callback_query.message.reply_text(
            "Ошибка проверки подписки. Убедись, что бот — админ канала."
        )
        return

    if status in ("creator", "administrator", "member"):
        await update.callback_query.message.reply_text(MSG_OK)
    else:
        await update.callback_query.message.reply_text(MSG_FAIL)

# Обработчик удаления/блокировки бота
async def my_chat_member(update, context: ContextTypes.DEFAULT_TYPE):
    cm = update.my_chat_member
    chat = cm.chat
    uid  = chat.id
    uname = chat.username or str(uid)
    new_status = cm.new_chat_member.status
    blocked = new_status in ('kicked', 'left')
    ts = now_moscow()

    logging.info(f"ChatMember update for user_id={uid}: new_status={new_status}")
    log_to_excel(ts, uid, uname, f"chat_member_{new_status}", False, blocked, 0)

# Точка входа
def main():
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(check_subscribed, pattern="^check_sub$"))
    app.add_handler(ChatMemberHandler(my_chat_member, ChatMemberHandler.MY_CHAT_MEMBER))
    app.run_polling()

if __name__ == "__main__":
    main()
